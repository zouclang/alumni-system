import openpyxl
wb = openpyxl.load_workbook('../大连理工大学苏州校友通讯录.xlsx', data_only=True)
ws = wb.active
for r in range(2, ws.max_row+1):
    val = ws.cell(row=r, column=2).value
    if val and '张鑫' in str(val):
        print('Row:', r, '张鑫:', ws.cell(row=r, column=5).value, '|', ws.cell(row=r, column=15).value)
        break
