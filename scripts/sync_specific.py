import sqlite3
import os
from importlib.machinery import SourceFileLoader

# Load the import.py logic so we can use its parse_school_experiences, clean_value, etc.
import_module = SourceFileLoader("import_script", "scripts/import.py").load_module()

EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), '..', '大连理工大学苏州校友通讯录.xlsx')
DB_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'alumni.db')

names_to_update = [
    "张鑫", "何书辉", "李之考", "石茂林", "王贇栋", "潘思敏", "王路征", "陈瀚泽", "陈梓澳", 
    "葛天祺", "路瑞琪", "王博艺", "杨小杨", "尹辉平", "张婉婷", "刘兆翔", "薄祥来", "冯李民", 
    "顾剑", "黄集长", "钱佳伟", "朱晓锋", "曹超杰", "李书兴", "马合建", "毛柘藤", "宋勇成", 
    "苏常慧", "谢晓军", "姚泓如", "袁向铭", "翟豪瑞", "李燕寒", "陈小贝", "沈奇"
]

def sync():
    import openpyxl
    print("Loading workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    count = 0
    # Find headers (row 1 or 2 often) and start row
    # In import.py it's:
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, 25)]
        name = import_module.clean_value(row[1])
        if name and import_module.re.sub(r'\d+$', '', name) in names_to_update:
            # We found a match in Excel
            clean_name = import_module.re.sub(r'\d+$', '', name)
            
            # What do we want to update?
            # They said "之前把所在微信群的内容错误的贴到在校经历里面了"
            # Now "我在通讯录原始表格里更新了这些人的信息"
            # So we extract the new school_experience (row[4]) and wechat_groups (row[14])
            
            new_exp_str = import_module.clean_value(row[4])
            new_wechat_group_str = import_module.clean_value(row[14])
            if new_wechat_group_str:
                new_wechat_group_str = str(new_wechat_group_str).replace('、', ',')
                
            college_norm_str = import_module.clean_value(row[8])

            # Update alumni table
            cur.execute("""
                UPDATE alumni 
                SET school_experience = ?, wechat_groups = ?
                WHERE name = ?
            """, (new_exp_str, new_wechat_group_str, clean_name))

            # Need to re-build school_experiences table for them
            # First find their alumni.id
            cur.execute("SELECT id FROM alumni WHERE name = ?", (clean_name,))
            res = cur.fetchall()
            for r_db in res:
                aid = r_db[0]
                cur.execute("DELETE FROM school_experiences WHERE alumni_id = ?", (aid,))
                experiences = import_module.parse_school_experiences(new_exp_str, college_norm_str)
                for exp in experiences:
                    cur.execute("""
                        INSERT INTO school_experiences (alumni_id, stage, start_year, end_year, college, sort_order)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (
                        aid,
                        exp.get('stage'),
                        exp.get('start_year'),
                        exp.get('end_year'),
                        exp.get('college'),
                        exp.get('sort_order')
                    ))
            count += 1

    conn.commit()
    conn.close()
    print(f"Updated {count} records successfully.")

if __name__ == '__main__':
    sync()
